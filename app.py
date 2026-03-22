
import io
import os
import re
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from streamlit_autorefresh import st_autorefresh


st.set_page_config(
    page_title="Nihaoma Education Center Dashboard",
    page_icon="🎓",
    layout="wide",
)

DEFAULT_REFRESH_SECONDS = int(os.getenv("REFRESH_SECONDS", "60"))
DEFAULT_DRIVE_LINK = os.getenv("GOOGLE_DRIVE_URL", "")
DEFAULT_DRIVE_FILE_ID = os.getenv("GOOGLE_DRIVE_FILE_ID", "")


def format_idr(value):
    try:
        return f"Rp {float(value):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


def extract_drive_file_id(value: str) -> str:
    if not value:
        return ""
    value = value.strip()
    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"id=([a-zA-Z0-9_-]+)",
        r"/d/([a-zA-Z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", value):
        return value
    return ""


@st.cache_data(ttl=60, show_spinner=False)
def download_excel_from_drive(file_id: str) -> bytes:
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    session = requests.Session()
    response = session.get(url, timeout=60)
    response.raise_for_status()

    confirm_token = None
    for key, val in response.cookies.items():
        if key.startswith("download_warning"):
            confirm_token = val
            break

    if confirm_token:
        response = session.get(url, params={"confirm": confirm_token}, timeout=60)
        response.raise_for_status()

    content_type = response.headers.get("content-type", "").lower()
    if "text/html" in content_type and len(response.content) < 50000:
        raise ValueError(
            "Google Drive mengembalikan halaman HTML, bukan file Excel. Pastikan file dibagikan "
            "dengan akses 'Anyone with the link' minimal Viewer."
        )
    return response.content


def parse_workbook(excel_bytes: bytes):
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=False)

    ws_input = wb["INPUT_DATA"]
    ws_setup = wb["SETUP"] if "SETUP" in wb.sheetnames else None

    brand_profile = {
        "brand_name": ws_input["B5"].value or "Nihaoma Education Center",
        "address": ws_input["B6"].value or "",
        "email": ws_input["B7"].value or "",
        "phone": ws_input["B8"].value or "",
        "payment_info": ws_input["B9"].value or "",
        "footer_note": ws_input["B10"].value or "",
        "default_due_days": ws_input["B11"].value or 7,
    }

    program_prices = {}
    if ws_setup:
        row = 2
        while True:
            program = ws_setup[f"A{row}"].value
            price = ws_setup[f"B{row}"].value
            if not program:
                break
            program_prices[str(program).strip()] = float(price or 0)
            row += 1

    rows = []
    r = 15
    while r <= ws_input.max_row:
        student_name = ws_input[f"D{r}"].value
        program = ws_input[f"H{r}"].value
        paid = ws_input[f"J{r}"].value
        if not student_name and not program and not paid:
            r += 1
            continue

        no_value = ws_input[f"A{r}"].value
        invoice_code = ws_input[f"B{r}"].value
        input_date = ws_input[f"C{r}"].value
        student_id = ws_input[f"E{r}"].value
        email = ws_input[f"F{r}"].value
        whatsapp = ws_input[f"G{r}"].value
        price = ws_input[f"I{r}"].value
        paid = ws_input[f"J{r}"].value or 0
        balance = ws_input[f"K{r}"].value
        payment_status = ws_input[f"L{r}"].value
        send_status = ws_input[f"M{r}"].value
        sent_date = ws_input[f"N{r}"].value
        notes = ws_input[f"O{r}"].value

        if no_value in (None, ""):
            no_value = r - 14
        if not invoice_code:
            invoice_code = f"NHEC-{int(no_value):04d}"

        if price in (None, "") and program:
            price = program_prices.get(str(program).strip(), 0)
        price = float(price or 0)
        paid = float(paid or 0)

        if balance in (None, ""):
            balance = max(price - paid, 0)
        balance = float(balance or 0)

        if not payment_status:
            payment_status = "Lunas" if balance <= 0 else "Belum Lunas"
        if not send_status:
            send_status = "Belum Dikirim"

        rows.append(
            {
                "no": int(no_value),
                "invoice_code": str(invoice_code),
                "input_date": input_date,
                "student_name": student_name or "",
                "student_id": student_id or "",
                "email": email or "",
                "whatsapp": whatsapp or "",
                "program": program or "",
                "price": price,
                "paid": paid,
                "balance": balance,
                "payment_status": payment_status,
                "send_status": send_status,
                "sent_date": sent_date,
                "notes": notes or "",
            }
        )
        r += 1

    df = pd.DataFrame(rows)

    if not df.empty:
        for col in ["input_date", "sent_date"]:
            df[col] = pd.to_datetime(df[col], errors="coerce")
        df["payment_status"] = df["payment_status"].fillna("Belum Lunas")
        df["send_status"] = df["send_status"].fillna("Belum Dikirim")
        df["is_lunas"] = df["payment_status"].str.lower().eq("lunas")
        df["is_sent"] = df["send_status"].str.lower().eq("sudah dikirim")
    else:
        df = pd.DataFrame(
            columns=[
                "no", "invoice_code", "input_date", "student_name", "student_id", "email",
                "whatsapp", "program", "price", "paid", "balance", "payment_status",
                "send_status", "sent_date", "notes", "is_lunas", "is_sent"
            ]
        )

    return df, brand_profile, program_prices


def load_source():
    upload = st.session_state.get("uploaded_excel")
    drive_source = st.session_state.get("drive_source", "").strip()
    file_id = extract_drive_file_id(drive_source) or DEFAULT_DRIVE_FILE_ID or extract_drive_file_id(DEFAULT_DRIVE_LINK)

    if upload is not None:
        return upload.getvalue(), "uploaded-file"

    if file_id:
        return download_excel_from_drive(file_id), f"google-drive:{file_id}"

    raise ValueError("Belum ada file Excel. Isi Google Drive link/file ID di sidebar atau upload file manual.")


def kpi_card(label, value, help_text=""):
    st.markdown(
        f"""
        <div style="padding:18px 18px 14px 18px;border-radius:20px;background:#ffffff;border:1px solid #ececf2;
                    box-shadow:0 8px 24px rgba(25,25,55,0.05);">
            <div style="font-size:0.9rem;color:#6b7280;margin-bottom:6px;">{label}</div>
            <div style="font-size:1.6rem;font-weight:700;color:#111827;">{value}</div>
            <div style="font-size:0.8rem;color:#9ca3af;margin-top:6px;">{help_text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


st.markdown(
    """
    <style>
    .block-container {padding-top: 1.4rem; padding-bottom: 1.5rem;}
    .stMetric {background: white; border-radius: 18px; padding: 10px;}
    .small-note {color:#6b7280;font-size:0.9rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("🎓 Nihaoma Education Center Dashboard")
st.caption("Dashboard web untuk memantau invoice student yang terhubung ke file Excel di Google Drive.")

with st.sidebar:
    st.header("Sumber Data")
    st.session_state["drive_source"] = st.text_input(
        "Google Drive link / file ID",
        value=st.session_state.get("drive_source", DEFAULT_DRIVE_LINK or DEFAULT_DRIVE_FILE_ID),
        help="Masukkan link share Google Drive atau file ID. File sebaiknya dibagikan sebagai 'Anyone with the link - Viewer'.",
    )
    st.file_uploader(
        "Atau upload Excel manual",
        type=["xlsx"],
        key="uploaded_excel",
        help="Dipakai hanya jika Anda ingin override sumber Google Drive sementara.",
    )
    refresh_seconds = st.slider(
        "Auto refresh (detik)",
        min_value=15,
        max_value=300,
        value=DEFAULT_REFRESH_SECONDS,
        step=15,
        help="Dashboard akan membaca ulang file dari Google Drive sesuai interval ini.",
    )
    st_autorefresh(interval=refresh_seconds * 1000, key="data_refresh")
    if st.button("Refresh sekarang", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.divider()
    st.subheader("Deploy ke web")
    st.markdown(
        """
        1. Upload project ini ke GitHub  
        2. Deploy di Streamlit Community Cloud  
        3. Isi `GOOGLE_DRIVE_URL` atau `GOOGLE_DRIVE_FILE_ID` di Secrets / Environment Variables  
        """
    )

try:
    excel_bytes, source_name = load_source()
    df, brand_profile, program_prices = parse_workbook(excel_bytes)
    last_refreshed = datetime.now().strftime("%d %b %Y %H:%M:%S")
except Exception as exc:
    st.error(f"Gagal memuat data: {exc}")
    st.stop()

left_header, right_header = st.columns([1.4, 1])
with left_header:
    st.markdown(
        f"""
        <div style="padding:18px;border-radius:22px;background:linear-gradient(135deg,#1f2b5c,#3747a4);color:white;">
            <div style="font-size:0.95rem;opacity:0.85;">Sumber data aktif</div>
            <div style="font-size:1.2rem;font-weight:700;margin-top:4px;">{brand_profile["brand_name"]}</div>
            <div style="margin-top:8px;opacity:0.9;">{source_name}</div>
            <div style="margin-top:6px;font-size:0.9rem;opacity:0.8;">Last refresh: {last_refreshed}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
with right_header:
    st.markdown(
        f"""
        <div style="padding:18px;border-radius:22px;background:#fff;border:1px solid #ececf2;">
            <div style="font-size:0.9rem;color:#6b7280;">Kontak Admin</div>
            <div style="font-weight:700;margin-top:6px;">{brand_profile["email"]}</div>
            <div style="margin-top:4px;color:#374151;">{brand_profile["phone"]}</div>
            <div style="margin-top:8px;color:#6b7280;font-size:0.9rem;">Auto refresh membaca ulang file Drive agar perubahan terbaru ikut tampil.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

if df.empty:
    st.warning("Data invoice belum ada di workbook.")
    st.stop()

filter_col1, filter_col2, filter_col3, filter_col4 = st.columns([1.1, 1, 1, 1.2])
with filter_col1:
    program_options = ["Semua Program"] + sorted([p for p in df["program"].dropna().unique().tolist() if p])
    selected_program = st.selectbox("Program", program_options)
with filter_col2:
    payment_options = ["Semua Status"] + sorted(df["payment_status"].dropna().unique().tolist())
    selected_payment = st.selectbox("Pelunasan", payment_options)
with filter_col3:
    send_options = ["Semua Status"] + sorted(df["send_status"].dropna().unique().tolist())
    selected_send = st.selectbox("Pengiriman", send_options)
with filter_col4:
    query = st.text_input("Cari student / kode invoice", placeholder="Contoh: Aulia / NHEC-0001")

filtered = df.copy()
if selected_program != "Semua Program":
    filtered = filtered[filtered["program"] == selected_program]
if selected_payment != "Semua Status":
    filtered = filtered[filtered["payment_status"] == selected_payment]
if selected_send != "Semua Status":
    filtered = filtered[filtered["send_status"] == selected_send]
if query.strip():
    q = query.strip().lower()
    filtered = filtered[
        filtered["student_name"].str.lower().str.contains(q, na=False)
        | filtered["invoice_code"].str.lower().str.contains(q, na=False)
        | filtered["email"].str.lower().str.contains(q, na=False)
    ]

total_students = int(filtered["student_name"].nunique())
total_tagihan = float(filtered["price"].sum())
total_paid = float(filtered["paid"].sum())
total_balance = float(filtered["balance"].sum())
lunas_pct = (float(filtered["is_lunas"].mean()) * 100) if len(filtered) else 0
sent_pct = (float(filtered["is_sent"].mean()) * 100) if len(filtered) else 0

k1, k2, k3, k4, k5, k6 = st.columns(6)
with k1:
    kpi_card("Invoice aktif", f"{len(filtered):,}".replace(",", "."), "Jumlah baris invoice setelah filter")
with k2:
    kpi_card("Student unik", f"{total_students:,}".replace(",", "."), "Nama student unik")
with k3:
    kpi_card("Total tagihan", format_idr(total_tagihan), "Total nilai invoice")
with k4:
    kpi_card("Sudah dibayar", format_idr(total_paid), "Akumulasi pembayaran")
with k5:
    kpi_card("Outstanding", format_idr(total_balance), "Sisa tagihan")
with k6:
    kpi_card("Lunas / Terkirim", f"{lunas_pct:.0f}% / {sent_pct:.0f}%", "Persentase dari hasil filter")

tab1, tab2, tab3, tab4 = st.tabs(["Overview", "Daftar Invoice", "Detail Invoice", "Data Source"])

with tab1:
    c1, c2 = st.columns([1.15, 1])
    with c1:
        program_summary = (
            filtered.groupby("program", dropna=False)[["price", "paid", "balance"]]
            .sum()
            .sort_values("price", ascending=False)
            .reset_index()
        )
        st.subheader("Ringkasan per Program")
        st.dataframe(
            program_summary.rename(
                columns={"program": "Program", "price": "Total Tagihan", "paid": "Sudah Dibayar", "balance": "Outstanding"}
            ),
            use_container_width=True,
            hide_index=True,
        )
    with c2:
        status_summary = pd.DataFrame(
            {
                "Kategori": ["Lunas", "Belum Lunas", "Sudah Dikirim", "Belum Dikirim"],
                "Jumlah": [
                    int((filtered["payment_status"] == "Lunas").sum()),
                    int((filtered["payment_status"] != "Lunas").sum()),
                    int((filtered["send_status"] == "Sudah Dikirim").sum()),
                    int((filtered["send_status"] != "Sudah Dikirim").sum()),
                ],
            }
        )
        st.subheader("Status Invoice")
        st.bar_chart(status_summary.set_index("Kategori"))

    st.subheader("10 Invoice Terbaru")
    latest = filtered.sort_values("input_date", ascending=False).head(10).copy()
    latest["input_date"] = latest["input_date"].dt.strftime("%d-%m-%Y")
    st.dataframe(
        latest[
            ["invoice_code", "input_date", "student_name", "program", "price", "paid", "balance", "payment_status", "send_status"]
        ].rename(
            columns={
                "invoice_code": "Kode Invoice",
                "input_date": "Tanggal",
                "student_name": "Student",
                "program": "Program",
                "price": "Tagihan",
                "paid": "Dibayar",
                "balance": "Sisa",
                "payment_status": "Pelunasan",
                "send_status": "Pengiriman",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

with tab2:
    st.subheader("Daftar Invoice Lengkap")
    table_df = filtered.copy()
    for col in ["input_date", "sent_date"]:
        table_df[col] = table_df[col].dt.strftime("%d-%m-%Y")
    st.dataframe(
        table_df[
            [
                "invoice_code", "input_date", "student_name", "student_id", "email",
                "whatsapp", "program", "price", "paid", "balance",
                "payment_status", "send_status", "sent_date", "notes"
            ]
        ].rename(
            columns={
                "invoice_code": "Kode Invoice",
                "input_date": "Tanggal Input",
                "student_name": "Nama Student",
                "student_id": "Passport / ID",
                "email": "Email",
                "whatsapp": "WhatsApp",
                "program": "Program",
                "price": "Harga Program",
                "paid": "Sudah Dibayar",
                "balance": "Sisa Tagihan",
                "payment_status": "Status Pelunasan",
                "send_status": "Status Pengiriman",
                "sent_date": "Tanggal Kirim",
                "notes": "Keterangan",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    csv = filtered.to_csv(index=False).encode("utf-8-sig")
    st.download_button("Download CSV hasil filter", data=csv, file_name="nihaoma_dashboard_filtered.csv", mime="text/csv")

with tab3:
    invoice_options = filtered["invoice_code"].tolist()
    default_idx = 0
    selected_invoice = st.selectbox("Pilih invoice", invoice_options, index=default_idx if invoice_options else None)
    chosen = filtered[filtered["invoice_code"] == selected_invoice].iloc[0]

    d1, d2 = st.columns([1.05, 0.95])
    with d1:
        st.markdown(
            f"""
            <div style="padding:22px;border-radius:24px;background:#fff;border:1px solid #ececf2;box-shadow:0 8px 24px rgba(25,25,55,0.05);">
                <div style="font-size:0.9rem;color:#6b7280;">Invoice</div>
                <div style="font-size:1.6rem;font-weight:700;color:#111827;">{chosen['invoice_code']}</div>
                <hr style="margin:14px 0;border:0;border-top:1px solid #eef2f7;">
                <div><b>Nama Student:</b> {chosen['student_name']}</div>
                <div><b>Passport / ID:</b> {chosen['student_id']}</div>
                <div><b>Email:</b> {chosen['email']}</div>
                <div><b>WhatsApp:</b> {chosen['whatsapp']}</div>
                <div><b>Program:</b> {chosen['program']}</div>
                <div><b>Tanggal Input:</b> {chosen['input_date'].strftime("%d %B %Y") if pd.notna(chosen['input_date']) else "-"}</div>
                <div><b>Keterangan:</b> {chosen['notes'] or "-"}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with d2:
        st.markdown(
            f"""
            <div style="padding:22px;border-radius:24px;background:linear-gradient(135deg,#1f2b5c,#3747a4);color:white;">
                <div style="font-size:0.9rem;opacity:0.85;">Ringkasan Pembayaran</div>
                <div style="margin-top:14px;"><b>Total Program:</b> {format_idr(chosen['price'])}</div>
                <div style="margin-top:10px;"><b>Sudah Dibayar:</b> {format_idr(chosen['paid'])}</div>
                <div style="margin-top:10px;"><b>Sisa Tagihan:</b> {format_idr(chosen['balance'])}</div>
                <div style="margin-top:10px;"><b>Status Pelunasan:</b> {chosen['payment_status']}</div>
                <div style="margin-top:10px;"><b>Status Pengiriman:</b> {chosen['send_status']}</div>
                <div style="margin-top:10px;"><b>Tanggal Kirim:</b> {chosen['sent_date'].strftime("%d %B %Y") if pd.notna(chosen['sent_date']) else "-"}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

with tab4:
    st.subheader("Informasi Sumber Data")
    st.json(
        {
            "brand_profile": brand_profile,
            "program_prices": program_prices,
            "active_source": source_name,
            "row_count": int(len(df)),
            "refresh_seconds": refresh_seconds,
        }
    )
    st.download_button(
        "Download workbook aktif",
        data=excel_bytes,
        file_name="nihaoma_source.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
